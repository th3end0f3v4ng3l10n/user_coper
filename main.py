#!/bin/env python3
from telethon import TelegramClient, sync
import csv
import configparser
from openpyxl import Workbook, load_workbook


class main():
    def __init__(self, ):
        """ИНИЦИАЛИЗАЦИЯ"""

        #Настройки
        self.config = configparser.ConfigParser() 
        self.config.read('settings.ini')
        self.api_id = self.config['Telegram']['api_id']
        self.api_hash = self.config['Telegram']['api_hash']
        self.phone = self.config['Telegram']['phone']
        
        #Ссылк на каналы
        self.urls = [
            'https://t.me/mpstatsio',
            'https://t.me/MarketplaceChati',
            'https://t.me/sellerfox',
            'https://t.me/joinchat/AAAAAEpHBwpyT6Z_cmG93g',
            'https://t.me/chat_marketplace ',
            'https://t.me/wildberries_otzivi',
            'https://t.me/prodavecwb',
            'https://t.me/marketguruclub']


        #Инициализация телеграм клиента
        self.client = TelegramClient(self.phone, self.api_id, self.api_hash) 

    def start(self):
        wb = load_workbook('output.xlsx')
        sheet = wb['Sheet1']
        self.client.start() #Запуск телеграм клиента
        h = 1
        for url in self.urls:
            try:
                print('\n', '[URL] ', url)
                participants = self.client.get_participants(url)
                for i in participants:
                # print(i,'\n','\n')
                    print(i.username,i.phone,'\n')
                    if (i.username != 'None') & (i.phone != 'None'):
                        h +=1
                        cell_obj = sheet.cell(row=h, column=1).value = i.username
                        cell_obj = sheet.cell(row=h, column=2).value = i.phone
                    
            except:
                print('Access Denied ', url)
                continue
        wb.save('output.xlsx')
 
root = main()
root.start()